"""
将后台数据导出为 Excel 表格，供甲方查阅。
生成两个 Sheet：
  Sheet1 - 总览：四六级评分 + 六维度汇总
  Sheet2 - 详细记录：每条子描述语的维度/子维度/描述语/得分
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 样式常量 ─────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="2F5496")   # 深蓝
_SECTION_FILL  = PatternFill("solid", fgColor="D6E4F7")   # 浅蓝
_ALT_FILL      = PatternFill("solid", fgColor="F2F7FF")   # 交替行浅色
_WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FONT   = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
_TITLE_FONT    = Font(name="微软雅黑", bold=True, size=12)
_BODY_FONT     = Font(name="微软雅黑", size=10)
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_DIM_SECTION_COLORS = {
    "argumentation": "C5D9F1",
    "discourse":     "C6EFCE",
    "convention":    "FFEB9C",
    "vocabulary":    "FFC7CE",
    "grammar":       "E2EFDA",
    "syntax":        "FCE4D6",
}

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _cell(ws, row, col, value, font=None, fill=None, alignment=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font      = font
    if fill:      c.fill      = fill
    if alignment: c.alignment = alignment
    if border:    c.border    = _BORDER
    return c


def _score_fill(score, max_score=4):
    """根据得分比例返回填充色：绿 → 黄 → 红"""
    ratio = score / max_score if max_score else 0
    if ratio >= 0.75:
        return PatternFill("solid", fgColor="C6EFCE")   # 绿
    elif ratio >= 0.5:
        return PatternFill("solid", fgColor="FFEB9C")   # 黄
    else:
        return PatternFill("solid", fgColor="FFC7CE")   # 红


def _final_score_fill(score, max_score=5):
    ratio = score / max_score if max_score else 0
    if ratio >= 0.75:
        return PatternFill("solid", fgColor="C6EFCE")
    elif ratio >= 0.5:
        return PatternFill("solid", fgColor="FFEB9C")
    else:
        return PatternFill("solid", fgColor="FFC7CE")


# ============================================================
# Sheet 1：总览
# ============================================================
def _build_summary_sheet(ws, backend_data: dict):
    ws.title = "总览"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60

    row = 1

    # 标题
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="英语写作评分与诊断报告 — 后台总览")
    c.font      = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    c.alignment = _CENTER
    ws.row_dimensions[row].height = 30
    row += 1

    # ── 四六级总分 ────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="一、四六级总体评分")
    c.font = _TITLE_FONT; c.fill = PatternFill("solid", fgColor="2F5496")
    c.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    c.alignment = _LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    headers = ["评分标准", "得分", "档位", "评分说明"]
    for col, h in enumerate(headers, 1):
        _cell(ws, row, col, h, font=_HEADER_FONT,
              fill=PatternFill("solid", fgColor="4472C4"), alignment=_CENTER)
    ws.row_dimensions[row].height = 20
    row += 1

    for label, score_key, band_key, rat_key in [
        ("四级标准", "cet4_score", "cet4_band", "cet4_rationale"),
        ("六级标准", "cet6_score", "cet6_band", "cet6_rationale"),
    ]:
        score = backend_data[score_key]
        sfill = _final_score_fill(score, 15)
        _cell(ws, row, 1, label,                   font=_BODY_FONT, alignment=_CENTER)
        _cell(ws, row, 2, f"{score} / 15",         font=Font(name="微软雅黑", bold=True, size=10),
              fill=sfill, alignment=_CENTER)
        _cell(ws, row, 3, backend_data[band_key],  font=_BODY_FONT, alignment=_CENTER)
        _cell(ws, row, 4, backend_data[rat_key],   font=_BODY_FONT, alignment=_LEFT)
        ws.row_dimensions[row].height = 45
        row += 1

    # ── 六维度汇总 ────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="二、六维度诊断汇总")
    c.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496"); c.alignment = _LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    headers = ["维度", "最终得分（/5）", "强项", "弱项"]
    for col, h in enumerate(headers, 1):
        _cell(ws, row, col, h, font=_HEADER_FONT,
              fill=PatternFill("solid", fgColor="4472C4"), alignment=_CENTER)
    ws.row_dimensions[row].height = 20
    row += 1

    dim_order = [
        ("argumentation", "论证能力"),
        ("discourse",     "语篇能力"),
        ("convention",    "书写规范"),
        ("vocabulary",    "词汇"),
        ("grammar",       "语法"),
        ("syntax",        "句法"),
    ]
    for dim_key, dim_name in dim_order:
        dim = backend_data["dimensions"].get(dim_key)
        if not dim:
            continue
        fs = dim["final_score"]
        sfill = _final_score_fill(fs)
        _cell(ws, row, 1, dim_name,          font=Font(name="微软雅黑", bold=True, size=10),
              fill=PatternFill("solid", fgColor=_DIM_SECTION_COLORS[dim_key]), alignment=_CENTER)
        _cell(ws, row, 2, f"{fs:.1f}",       font=Font(name="微软雅黑", bold=True, size=10),
              fill=sfill, alignment=_CENTER)
        _cell(ws, row, 3, dim["strengths"],  font=_BODY_FONT, alignment=_LEFT)
        _cell(ws, row, 4, dim["weaknesses"], font=_BODY_FONT, alignment=_LEFT)
        ws.row_dimensions[row].height = 55
        row += 1


# ============================================================
# Sheet 2：子描述语详细记录
# ============================================================
def _build_detail_sheet(ws, backend_data: dict):
    ws.title = "详细评分记录"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 10

    row = 1

    # 标题
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="六维度子描述语详细评分记录")
    c.font      = Font(name="微软雅黑", bold=True, size=13, color="2F5496")
    c.alignment = _CENTER
    ws.row_dimensions[row].height = 28
    row += 1

    # 说明行
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1,
                value="说明：子描述语得分范围 0-4 分；最终维度得分 = (各项总分/项数)/4×5，范围 0-5 分")
    c.font = Font(name="微软雅黑", italic=True, size=9, color="666666")
    c.alignment = _LEFT
    ws.row_dimensions[row].height = 16
    row += 1

    dim_order = [
        ("argumentation", "论证能力"),
        ("discourse",     "语篇能力"),
        ("convention",    "书写规范"),
        ("vocabulary",    "词汇"),
        ("grammar",       "语法"),
        ("syntax",        "句法"),
    ]

    for dim_key, dim_name in dim_order:
        dim = backend_data["dimensions"].get(dim_key)
        if not dim:
            continue
        sec_color = _DIM_SECTION_COLORS[dim_key]

        # 维度分组标题行
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1,
                    value=f"{dim_name}   最终得分：{dim['final_score']:.1f} / 5.0")
        c.font      = Font(name="微软雅黑", bold=True, size=11)
        c.fill      = PatternFill("solid", fgColor=sec_color)
        c.alignment = _LEFT; c.border = _BORDER
        ws.row_dimensions[row].height = 22
        row += 1

        # 列头
        has_subdim = "子维度" in (dim["detail_table"][0] if dim["detail_table"] else {})
        if has_subdim:
            col_headers = ["维度", "子维度", "描述语", "得分"]
        else:
            col_headers = ["维度", "描述语", "得分", ""]
        for col, h in enumerate(col_headers, 1):
            _cell(ws, row, col, h, font=_HEADER_FONT,
                  fill=PatternFill("solid", fgColor="4472C4"), alignment=_CENTER)
        ws.row_dimensions[row].height = 18
        row += 1

        # 数据行
        for i, record in enumerate(dim["detail_table"]):
            fill = PatternFill("solid", fgColor=sec_color) if i % 2 == 0 else _WHITE_FILL
            score = record.get("得分")
            sfill = _score_fill(score) if score is not None else fill

            if has_subdim:
                _cell(ws, row, 1, record.get("维度", ""),   font=_BODY_FONT, fill=fill, alignment=_CENTER)
                _cell(ws, row, 2, record.get("子维度", ""), font=_BODY_FONT, fill=fill, alignment=_CENTER)
                _cell(ws, row, 3, record.get("描述语", ""), font=_BODY_FONT, fill=fill, alignment=_LEFT)
                _cell(ws, row, 4, score,                    font=Font(name="微软雅黑", bold=True, size=10),
                      fill=sfill, alignment=_CENTER)
            else:
                _cell(ws, row, 1, record.get("维度", ""),   font=_BODY_FONT, fill=fill, alignment=_CENTER)
                _cell(ws, row, 2, record.get("描述语", ""), font=_BODY_FONT, fill=fill, alignment=_LEFT)
                _cell(ws, row, 3, score,                    font=Font(name="微软雅黑", bold=True, size=10),
                      fill=sfill, alignment=_CENTER)
                _cell(ws, row, 4, "",                       fill=fill)
            ws.row_dimensions[row].height = 38
            row += 1


# ============================================================
# 主导出函数
# ============================================================
def export_to_excel(backend_data: dict, filepath: str = "report.xlsx"):
    """将后台数据导出为格式化 Excel 文件。"""
    wb = openpyxl.Workbook()

    # 默认 sheet 改造为总览
    ws_summary = wb.active
    _build_summary_sheet(ws_summary, backend_data)

    # 新增详细记录 sheet
    ws_detail = wb.create_sheet()
    _build_detail_sheet(ws_detail, backend_data)

    wb.save(filepath)
    print(f"Excel 报告已保存至 {filepath}")
